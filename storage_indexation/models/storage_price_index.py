# -*- coding: utf-8 -*-

import logging
import requests
import io
from datetime import date, datetime
from dateutil.relativedelta import relativedelta

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

# URL fichier Excel Statbel pour l'indice des prix à la consommation et indice santé
STATBEL_XLSX_URL = (
    "https://statbel.fgov.be/sites/default/files/files/opendata/"
    "Consumptieprijsindex%20en%20gezondheidsindex/CPI%20All%20base%20years.xlsx"
)

# Année de base utilisée pour les indices Lolirine.
# Statbel publie plusieurs bases simultanément (1914, 1988, 1996, 2004, 2013, 2025).
# La base 2025=100 a été introduite en janvier 2026 et remplace la base 2013.
TARGET_BASE_YEAR = 2025

# Bornes de validation pour rejeter les valeurs aberrantes
# (typiquement la colonne MS_CPI_INFL qui contient une variation en %, ex: 0.04)
MIN_VALID_INDEX_VALUE = 50.0
MAX_VALID_INDEX_VALUE = 200.0


class StoragePriceIndex(models.Model):
    """Modèle pour stocker les indices de prix (indice santé, CPI, etc.)"""
    _name = 'storage.price.index'
    _description = 'Indice de prix pour indexation'
    _order = 'date desc, id desc'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    name = fields.Char(
        string='Référence',
        required=True,
        tracking=True,
        help="Référence unique de l'indice (ex: 2026-04)"
    )
    index_type = fields.Selection([
        ('health', 'Indice Santé Belge'),
        ('cpi', 'Indice des Prix à la Consommation (CPI)'),
        ('custom', 'Indice Personnalisé'),
    ], string="Type d'indice", required=True, default='health', tracking=True)

    date = fields.Date(
        string='Date',
        required=True,
        tracking=True,
        help="Date de référence de l'indice (généralement le 1er du mois)"
    )
    year = fields.Integer(
        string='Année',
        compute='_compute_year_month',
        store=True
    )
    month = fields.Integer(
        string='Mois',
        compute='_compute_year_month',
        store=True
    )

    value = fields.Float(
        string="Valeur de l'indice",
        required=True,
        digits=(10, 2),
        tracking=True,
        help="Valeur numérique de l'indice"
    )
    base_year = fields.Integer(
        string='Année de base',
        default=TARGET_BASE_YEAR,
        help="Année de référence pour le calcul de l'indice (base 100). "
             "Statbel a introduit la base 2025 en janvier 2026, en remplacement "
             "de la base 2013."
    )

    source = fields.Selection([
        ('statbel', 'Statbel (automatique)'),
        ('manual', 'Saisie manuelle'),
        ('import', 'Import fichier'),
    ], string='Source', default='manual', tracking=True)

    active = fields.Boolean(default=True)
    company_id = fields.Many2one(
        'res.company',
        string='Société',
        default=lambda self: self.env.company
    )

    notes = fields.Text(string='Notes')

    unique_index_date_type = models.Constraint(
        'UNIQUE(date, index_type, company_id)',
        "Un seul indice par date et type est autorisé par société!",
    )

    @api.depends('date')
    def _compute_year_month(self):
        for record in self:
            if record.date:
                record.year = record.date.year
                record.month = record.date.month
            else:
                record.year = 0
                record.month = 0

    # ========================================================================
    # SCRAPING STATBEL
    # ========================================================================

    @api.model
    def _fetch_statbel_health_index(self, year=None, month=None):
        """Récupère l'indice santé belge depuis le fichier Excel Statbel.

        Le fichier Statbel actuel ('TA_ALL_BASE_YR') contient plusieurs lignes
        par mois, une par année de base (1914, 1988, 1996, 2004, 2013, 2025).
        On filtre sur TARGET_BASE_YEAR pour ne récupérer que les valeurs
        en base 2025.

        En-têtes attendus :
          - NM_YR        : année
          - NM_MTH       : mois
          - MS_CPI_IDX   : indice prix conso
          - MS_CPI_INFL  : variation inflation (% en décimal — NE PAS UTILISER)
          - MS_HLTH_IDX  : indice santé ← C'EST CE QU'ON VEUT
          - NM_BASE_YR   : année de base de la ligne

        Args:
            year: année cible (optionnel, sinon dernière dispo)
            month: mois cible (optionnel, sinon dernier dispo)

        Returns:
            dict {value, date, source, base_year} ou raise UserError
        """
        try:
            _logger.info(
                "Récupération indice santé Statbel (base %s)...",
                TARGET_BASE_YEAR
            )

            # ----------------------------------------------------------------
            # 1. Téléchargement
            # ----------------------------------------------------------------
            response = requests.get(STATBEL_XLSX_URL, timeout=60)
            response.raise_for_status()

            try:
                import openpyxl
            except ImportError:
                raise UserError(_(
                    "Le module 'openpyxl' est requis pour lire les fichiers Excel. "
                    "Veuillez l'installer avec: pip install openpyxl"
                ))

            wb = openpyxl.load_workbook(
                io.BytesIO(response.content),
                data_only=True
            )

            # ----------------------------------------------------------------
            # 2. Sélection de la feuille
            # ----------------------------------------------------------------
            sheet = None
            for sheet_name in wb.sheetnames:
                if 'BASE_YR' in sheet_name.upper():
                    sheet = wb[sheet_name]
                    break
            if not sheet:
                # Fallback : essai par mots-clés "santé / health" comme avant
                for sheet_name in wb.sheetnames:
                    sn_low = sheet_name.lower()
                    if 'health' in sn_low or 'santé' in sn_low or 'gezondheid' in sn_low:
                        sheet = wb[sheet_name]
                        break
            if not sheet:
                sheet = wb.active

            _logger.info(
                "Lecture de la feuille: %s (%d lignes)",
                sheet.title, sheet.max_row
            )

            # ----------------------------------------------------------------
            # 3. Identification des colonnes par codes Statbel exacts
            # ----------------------------------------------------------------
            headers = {}
            for col in range(1, sheet.max_column + 1):
                v = sheet.cell(row=1, column=col).value
                if v:
                    headers[str(v).strip()] = col

            required = ['NM_YR', 'NM_MTH', 'MS_HLTH_IDX', 'NM_BASE_YR']
            missing = [h for h in required if h not in headers]
            if missing:
                raise UserError(_(
                    "Format Statbel inattendu. En-têtes manquants : %s\n"
                    "En-têtes trouvés : %s\n\n"
                    "Statbel a peut-être modifié à nouveau le format du fichier. "
                    "Veuillez contacter le support technique."
                ) % (', '.join(missing), ', '.join(headers.keys())))

            year_col = headers['NM_YR']
            month_col = headers['NM_MTH']
            health_col = headers['MS_HLTH_IDX']
            base_col = headers['NM_BASE_YR']

            _logger.info(
                "Colonnes Statbel : NM_YR=%d, NM_MTH=%d, MS_HLTH_IDX=%d, "
                "NM_BASE_YR=%d",
                year_col, month_col, health_col, base_col
            )

            # ----------------------------------------------------------------
            # 4. Extraction des indices pour la base cible
            # ----------------------------------------------------------------
            indices = []
            rejected_aberrant = 0

            for row in range(2, sheet.max_row + 1):
                row_base = sheet.cell(row=row, column=base_col).value

                # Filtre principal : on ne garde QUE la base cible
                if row_base != TARGET_BASE_YEAR:
                    continue

                row_year = sheet.cell(row=row, column=year_col).value
                row_month = sheet.cell(row=row, column=month_col).value
                row_value = sheet.cell(row=row, column=health_col).value

                if not (row_year and row_month and row_value):
                    continue

                try:
                    value = float(row_value)
                except (ValueError, TypeError):
                    continue

                # GARDE-FOU : un indice santé en base 2025 est typiquement
                # entre 50 (années 90) et 200 (futur lointain).
                # Une valeur < 50 ou > 200 est probablement un % ou un bug.
                if not (MIN_VALID_INDEX_VALUE <= value <= MAX_VALID_INDEX_VALUE):
                    rejected_aberrant += 1
                    _logger.warning(
                        "Valeur aberrante ignorée pour %s-%s base %s : %s",
                        row_year, row_month, row_base, value
                    )
                    continue

                indices.append({
                    'year': int(row_year),
                    'month': int(row_month),
                    'value': value,
                })

            wb.close()

            if not indices:
                raise UserError(_(
                    "Aucun indice santé base %s trouvé dans le fichier Statbel."
                ) % TARGET_BASE_YEAR)

            # Trier par date décroissante (plus récent en tête)
            indices.sort(key=lambda x: (x['year'], x['month']), reverse=True)

            _logger.info(
                "Indices base %s extraits : %d (du %s-%02d au %s-%02d)"
                "%s",
                TARGET_BASE_YEAR, len(indices),
                indices[-1]['year'], indices[-1]['month'],
                indices[0]['year'], indices[0]['month'],
                f" — {rejected_aberrant} valeur(s) aberrante(s) rejetée(s)"
                    if rejected_aberrant else "",
            )

            # ----------------------------------------------------------------
            # 5. Filtrage par année/mois demandé
            # ----------------------------------------------------------------
            if year and month:
                for idx in indices:
                    if idx['year'] == int(year) and idx['month'] == int(month):
                        return {
                            'value': idx['value'],
                            'date': date(idx['year'], idx['month'], 1),
                            'source': 'statbel',
                            'base_year': TARGET_BASE_YEAR,
                        }
                raise UserError(_(
                    "Indice non trouvé pour %s/%s en base %s. "
                    "Dernier disponible : %s/%s = %s"
                ) % (
                    month, year, TARGET_BASE_YEAR,
                    indices[0]['month'], indices[0]['year'], indices[0]['value']
                ))

            # Retourner le plus récent
            latest = indices[0]
            return {
                'value': latest['value'],
                'date': date(latest['year'], latest['month'], 1),
                'source': 'statbel',
                'base_year': TARGET_BASE_YEAR,
            }

        except requests.RequestException as e:
            _logger.error("Erreur réseau récupération Statbel : %s", e)
            raise UserError(_(
                "Impossible de récupérer l'indice depuis Statbel. "
                "Veuillez réessayer plus tard ou saisir l'indice manuellement.\n\n"
                "Erreur technique : %s"
            ) % str(e))
        except UserError:
            # Laisser passer les erreurs métier (formatées proprement)
            raise
        except Exception as e:
            _logger.exception("Erreur inattendue récupération Statbel")
            raise UserError(_("Erreur inattendue : %s") % str(e))

    # ========================================================================
    # CRÉATION / MISE À JOUR
    # ========================================================================

    @api.model
    def fetch_latest_health_index(self):
        """Récupère et enregistre le dernier indice santé disponible.

        Si un indice existe déjà pour ce mois/société, il est mis à jour.
        Sinon, un nouveau record est créé.

        Returns:
            recordset storage.price.index (le record créé ou mis à jour)
            False si l'API n'a rien retourné
        """
        result = self._fetch_statbel_health_index()

        if not result:
            return False

        # Vérifier si l'indice existe déjà pour ce mois et cette société
        existing = self.search([
            ('date', '=', result['date']),
            ('index_type', '=', 'health'),
            ('company_id', '=', self.env.company.id),
        ], limit=1)

        vals = {
            'value': result['value'],
            'source': 'statbel',
            'base_year': result.get('base_year', TARGET_BASE_YEAR),
        }

        if existing:
            existing.write(vals)
            _logger.info(
                "Indice santé mis à jour : %s = %s (base %s, société %s)",
                result['date'], result['value'],
                vals['base_year'], self.env.company.name,
            )
            return existing

        # Création nouveau record
        vals.update({
            'name': f"Indice Santé {result['date'].strftime('%Y-%m')}",
            'index_type': 'health',
            'date': result['date'],
        })
        new_record = self.create(vals)
        _logger.info(
            "Indice santé créé : %s = %s (base %s, société %s)",
            result['date'], result['value'],
            vals['base_year'], self.env.company.name,
        )
        return new_record

    # ========================================================================
    # ACCESSEURS UTILITAIRES
    # ========================================================================

    @api.model
    def get_index_for_date(self, target_date, index_type='health'):
        """Récupère l'indice pour une date donnée.

        Si pas d'indice exact pour cette date, retourne le plus proche
        précédent (le plus récent <= target_date).
        """
        domain = [
            ('index_type', '=', index_type),
            ('date', '<=', target_date),
            ('company_id', '=', self.env.company.id),
        ]
        return self.search(domain, order='date desc', limit=1)

    @api.model
    def get_base_index(self, contract_date, index_type='health'):
        """Récupère l'indice de base pour un contrat (date de signature)."""
        return self.get_index_for_date(contract_date, index_type)

    @api.model
    def get_current_index(self, index_type='health'):
        """Récupère l'indice le plus récent."""
        return self.get_index_for_date(date.today(), index_type)

    # ========================================================================
    # ACTIONS UI / CRON
    # ========================================================================

    def action_fetch_from_statbel(self):
        """Bouton pour récupérer/mettre à jour l'indice depuis Statbel."""
        self.ensure_one()
        if self.index_type != 'health':
            raise UserError(_(
                "Seul l'indice santé peut être récupéré depuis Statbel"
            ))

        result = self._fetch_statbel_health_index(self.year, self.month)
        if result:
            self.write({
                'value': result['value'],
                'source': 'statbel',
                'base_year': result.get('base_year', TARGET_BASE_YEAR),
            })
        return True

    @api.model
    def _cron_fetch_health_index(self):
        """CRON : Récupère automatiquement l'indice santé chaque mois.

        Exécuté typiquement le 5 ou 10 du mois (selon configuration cron).
        Itère sur toutes les sociétés pour mettre à jour l'indice de chacune.
        """
        _logger.info("CRON: Récupération automatique de l'indice santé...")

        companies = self.env['res.company'].search([])
        for company in companies:
            try:
                self.with_company(company).fetch_latest_health_index()
                _logger.info(
                    "Indice santé mis à jour pour %s",
                    company.name
                )
            except Exception as e:
                _logger.error(
                    "Erreur mise à jour indice pour %s : %s",
                    company.name, e
                )
