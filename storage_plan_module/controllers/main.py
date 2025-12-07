# -*- coding: utf-8 -*-
from odoo import http
from odoo.http import request
import json
import io
import base64


class StoragePlanController(http.Controller):
    
    @http.route('/storage/plan', type='http', auth='public', website=True)
    def storage_plan(self, **kwargs):
        """Page principale du plan interactif"""
        # Récupérer tous les boxes actifs
        all_boxes = request.env['storage.box'].sudo().search([('active', '=', True)])
        
        # Créer un dictionnaire indexé par nom de box
        boxes_dict = {}
        for box in all_boxes:
            date_available_str = ''
            if box.date_available:
                date_available_str = box.date_available.strftime('%d/%m/%Y')
            
            boxes_dict[box.name] = {
                'id': box.id,
                'name': box.name,
                'surface': round(box.surface, 1),
                'volume': round(box.volume, 1),
                'status': box.status,
                'status_color': box.get_status_color(),
                'date_available': date_available_str,
                'floor': box.floor_id.name if box.floor_id else '',
            }
        
        # Légende des statuts depuis la configuration
        StatusColor = request.env['storage.status.color'].sudo()
        status_legend = StatusColor.get_legend_items()
        
        return request.render('storage_plan_module.storage_plan_page', {
            'boxes': boxes_dict,
            'status_legend': status_legend,
        })
    
    @http.route('/storage/boxes/export/xlsx', type='http', auth='user')
    def export_boxes_xlsx(self, **kwargs):
        """Exporte les boxes vers un fichier XLSX"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return request.make_response(
                "Module openpyxl non installé. Veuillez l'installer avec: pip install openpyxl",
                headers=[('Content-Type', 'text/plain')]
            )
        
        # Créer le workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Boxes"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = [
            'name', 'floor', 'floor_code', 'width', 'depth', 'height',
            'volume', 'surface', 'price_monthly', 'registration_fee',
            'deposit_months', 'status', 'date_available', 'aisle',
            'description', 'active'
        ]
        
        headers_labels = [
            'Nom du Box', 'Étage', 'Code Étage', 'Largeur (cm)', 'Profondeur (cm)', 'Hauteur (cm)',
            'Volume (m³)', 'Surface (m²)', 'Prix Mensuel (€)', 'Frais Dossier (€)',
            'Caution (mois)', 'Statut', 'Date Disponibilité', 'Allée',
            'Description', 'Actif'
        ]
        
        for col, (header, label) in enumerate(zip(headers, headers_labels), 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Données
        Box = request.env['storage.box'].sudo()
        data = Box.get_export_data()
        
        for row_num, row_data in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                if header == 'active':
                    value = 'Oui' if value else 'Non'
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
        
        # Ajuster les largeurs de colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Générer le fichier
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = "boxes_export.xlsx"
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"'),
            ]
        )
    
    @http.route('/storage/box/<int:box_id>/details', type='json', auth='public')
    def box_details(self, box_id, **kwargs):
        """Récupère les détails d'un box"""
        box = request.env['storage.box'].sudo().browse(box_id)
        if not box.exists():
            return {'error': 'Box non trouvé'}
        
        return box.get_box_details()
    
    @http.route('/storage/box/<int:box_id>/reserve', type='json', auth='public')
    def reserve_box(self, box_id, customer_name, customer_email, customer_phone, 
                    reservation_type='reservation', notes='', **kwargs):
        """Crée une réservation pour un box"""
        box = request.env['storage.box'].sudo().browse(box_id)
        if not box.exists():
            return {'error': 'Box non trouvé'}
        
        if box.status != 'disponible':
            return {'error': 'Ce box n\'est pas disponible'}
        
        try:
            reservation = request.env['box.reservation'].sudo().create({
                'box_id': box_id,
                'customer_name': customer_name,
                'customer_email': customer_email,
                'customer_phone': customer_phone,
                'reservation_type': reservation_type,
                'notes': notes,
                'state': 'pending',
            })
            
            # Mettre à jour le statut du box
            box.status = 'reserve'
            
            return {
                'success': True,
                'reservation_id': reservation.id,
                'reservation_ref': reservation.name,
                'message': 'Réservation créée avec succès'
            }
        except Exception as e:
            return {'error': str(e)}
    
    @http.route('/storage/box/<int:box_id>/appointment', type='json', auth='public')
    def book_appointment(self, box_id, customer_name, customer_email, customer_phone, 
                        appointment_date=None, notes='', **kwargs):
        """Crée une demande de rendez-vous pour un box"""
        box = request.env['storage.box'].sudo().browse(box_id)
        if not box.exists():
            return {'error': 'Box non trouvé'}
        
        try:
            reservation_vals = {
                'box_id': box_id,
                'customer_name': customer_name,
                'customer_email': customer_email,
                'customer_phone': customer_phone,
                'reservation_type': 'appointment',
                'notes': notes,
                'state': 'pending',
            }
            
            if appointment_date:
                reservation_vals['appointment_date'] = appointment_date
            
            reservation = request.env['box.reservation'].sudo().create(reservation_vals)
            
            return {
                'success': True,
                'reservation_id': reservation.id,
                'reservation_ref': reservation.name,
                'message': 'Demande de rendez-vous créée avec succès'
            }
        except Exception as e:
            return {'error': str(e)}
    
    @http.route('/storage/boxes/search', type='json', auth='public')
    def search_boxes(self, status=None, min_volume=None, max_volume=None, 
                     min_price=None, max_price=None, floor_id=None, **kwargs):
        """Recherche des boxes selon des critères"""
        domain = [('active', '=', True)]
        
        if status:
            domain.append(('status', '=', status))
        if min_volume:
            domain.append(('volume', '>=', float(min_volume)))
        if max_volume:
            domain.append(('volume', '<=', float(max_volume)))
        if min_price:
            domain.append(('price_monthly', '>=', float(min_price)))
        if max_price:
            domain.append(('price_monthly', '<=', float(max_price)))
        if floor_id:
            domain.append(('floor_id', '=', int(floor_id)))
        
        boxes = request.env['storage.box'].sudo().search(domain)
        
        results = []
        for box in boxes:
            results.append(box.get_box_details())
        
        return {
            'success': True,
            'boxes': results,
            'count': len(results)
        }
