# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io


class StorageBoxImportWizard(models.TransientModel):
    _name = 'storage.box.import.wizard'
    _description = 'Import de boxes depuis XLSX'
    
    file = fields.Binary(string='Fichier XLSX', required=True)
    filename = fields.Char(string='Nom du fichier')
    
    def action_import(self):
        """Importe les boxes depuis le fichier XLSX"""
        if not self.file:
            raise UserError(_("Veuillez sélectionner un fichier XLSX."))
        
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("Module openpyxl non installé. Veuillez l'installer avec: pip install openpyxl"))
        
        # Lire le fichier
        file_content = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        # Lire les en-têtes
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Mapping des labels vers les clés
        label_to_key = {
            'Nom du Box': 'name',
            'Étage': 'floor',
            'Code Étage': 'floor_code',
            'Largeur (cm)': 'width',
            'Profondeur (cm)': 'depth',
            'Hauteur (cm)': 'height',
            'Volume (m³)': 'volume',
            'Surface (m²)': 'surface',
            'Prix Mensuel (€)': 'price_monthly',
            'Frais Dossier (€)': 'registration_fee',
            'Caution (mois)': 'deposit_months',
            'Statut': 'status',
            'Date Disponibilité': 'date_available',
            'Allée': 'aisle',
            'Description': 'description',
            'Actif': 'active',
            # Support des clés anglaises aussi
            'name': 'name',
            'floor': 'floor',
            'floor_code': 'floor_code',
            'width': 'width',
            'depth': 'depth',
            'height': 'height',
            'price_monthly': 'price_monthly',
            'registration_fee': 'registration_fee',
            'deposit_months': 'deposit_months',
            'status': 'status',
            'date_available': 'date_available',
            'aisle': 'aisle',
            'description': 'description',
            'active': 'active',
        }
        
        # Convertir les headers en clés
        header_keys = []
        for h in headers:
            key = label_to_key.get(h, h) if h else None
            header_keys.append(key)
        
        # Lire les données
        data_rows = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(header_keys) and header_keys[col_idx]:
                    value = cell.value
                    # Convertir 'Oui'/'Non' en booléen
                    if header_keys[col_idx] == 'active':
                        value = value in [True, 'Oui', 'oui', 'OUI', 1, '1']
                    row_data[header_keys[col_idx]] = value
            data_rows.append(row_data)
        
        # Importer les données
        Box = self.env['storage.box']
        result = Box.import_from_xlsx_data(data_rows)
        
        # Message de résultat
        message = f"Import terminé !\n\n"
        message += f"✓ Boxes créés : {result['created']}\n"
        message += f"✓ Boxes mis à jour : {result['updated']}\n"
        
        if result['errors']:
            message += f"\n⚠ Erreurs ({len(result['errors'])}) :\n"
            for error in result['errors'][:10]:  # Limiter à 10 erreurs
                message += f"  - {error}\n"
            if len(result['errors']) > 10:
                message += f"  ... et {len(result['errors']) - 10} autres erreurs\n"
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import XLSX'),
                'message': message,
                'type': 'success' if not result['errors'] else 'warning',
                'sticky': True,
            }
        }
    
    def action_download_template(self):
        """Télécharge un fichier template XLSX"""
        return {
            'type': 'ir.actions.act_url',
            'url': '/storage/boxes/export/xlsx',
            'target': 'new',
        }
